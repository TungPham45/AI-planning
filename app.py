from flask import Flask, render_template, request, redirect, jsonify, send_from_directory
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime, timedelta
import base64
from werkzeug.utils import secure_filename
import time
from learning_path_ai import LearningPathAI
from generate_training_data import generate_training_data
import pandas as pd

app = Flask(__name__)
ai = LearningPathAI()

# Excel file constants
USERS_EXCEL = "users.xlsx"
LEARNING_PATH_EXCEL = "learning_path.xlsx"
DAILY_PLANS_EXCEL = "daily_plans.xlsx"
GRADE_EXCEL = "grade.xlsx"
SUBJECT_EXCEL = "subject.xlsx"
TOPIC_EXCEL = "topic.xlsx"
THEORY_EXCEL = "theory.xlsx"
PRACTICE_EXCEL = "practice.xlsx"

UPLOAD_FOLDER = 'static/uploads'
DEFAULT_AVATAR = 'static/default-avatar.svg'

# Tạo thư mục uploads nếu chưa tồn tại
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# Tạo ảnh mặc định nếu chưa tồn tại
if not os.path.exists(DEFAULT_AVATAR):
    # Tạo một ảnh SVG đơn giản
    svg_content = '''<?xml version="1.0" encoding="UTF-8" standalone="no"?>
    <svg width="200" height="200" xmlns="http://www.w3.org/2000/svg">
        <rect width="200" height="200" fill="#e0e0e0"/>
        <circle cx="100" cy="100" r="50" fill="#9e9e9e"/>
    </svg>'''
    
    # Lưu file SVG
    with open(DEFAULT_AVATAR, 'w') as f:
        f.write(svg_content)

def init_all_excel_files():
    """Khởi tạo tất cả các file Excel cần thiết"""
    excel_files = {
        USERS_EXCEL: {
            'sheet_name': 'Users',
            'headers': ['student_id', 'full_name', 'email', 'password', 'dob', 'join_date', 'image_url']
        },
        LEARNING_PATH_EXCEL: {
            'sheet_name': 'LearningPaths',
            'headers': ['path_id', 'user_id', 'subject', 'current_score', 'target_score', 
                       'duration_weeks', 'daily_study_hours', 'learning_style', 
                       'success_rate', 'start_date', 'end_date', 'status']
        },
        DAILY_PLANS_EXCEL: {
            'sheet_name': 'DailyPlans',
            'headers': ['plan_id', 'path_id', 'date', 'theory_topics', 'practice_exercises',
                       'theory_hours', 'practice_hours', 'learning_resources', 'completed']
        },
        GRADE_EXCEL: {
            'sheet_name': 'Grades',
            'headers': ['ID_grade', 'Name_grade']
        },
        SUBJECT_EXCEL: {
            'sheet_name': 'Subjects',
            'headers': ['ID_subject', 'name_subject', 'ID_grade']
        },
        TOPIC_EXCEL: {
            'sheet_name': 'Topics',
            'headers': ['ID_topic', 'topic_name', 'ID_subject', 'ID_grade']
        },
        THEORY_EXCEL: {
            'sheet_name': 'Theory',
            'headers': ['ID_theory', 'theory_name', 'ID_topic', 'ID_subject', 'ID_grade',
                       'level', 'URL', 'completion_time']
        },
        PRACTICE_EXCEL: {
            'sheet_name': 'Practice',
            'headers': ['ID_practice', 'practice_name', 'ID_topic', 'ID_subject', 'ID_grade',
                       'level', 'ID_theory']
        }
    }

    for file_name, config in excel_files.items():
        try:
            if not os.path.exists(file_name):
                wb = Workbook()
                ws = wb.active
                ws.title = config['sheet_name']
                ws.append(config['headers'])
                wb.save(file_name)
                print(f"Đã tạo file Excel mới: {file_name}")
            else:
                # Kiểm tra xem file có phải là file Excel hợp lệ không
                try:
                    wb = load_workbook(file_name)
                    print(f"File Excel đã tồn tại và hợp lệ: {file_name}")
                except Exception as e:
                    print(f"File Excel không hợp lệ, đang tạo lại: {e}")
                    os.remove(file_name)
                    wb = Workbook()
                    ws = wb.active
                    ws.title = config['sheet_name']
                    ws.append(config['headers'])
                    wb.save(file_name)
                    print(f"Đã tạo lại file Excel: {file_name}")
        except Exception as e:
            print(f"Lỗi khi khởi tạo Excel {file_name}: {e}")
            raise

# Gọi hàm khởi tạo Excel khi khởi động ứng dụng
init_all_excel_files()

# Kiểm tra xem mô hình đã được huấn luyện chưa
print("\nKiểm tra và khởi tạo mô hình AI...")
if os.path.exists('learning_path_model.joblib'):
    try:
        print("Tìm thấy file mô hình, đang tải...")
        ai.load_model()
        if not ai.is_trained:
            print("Mô hình chưa được huấn luyện, đang huấn luyện lại...")
            # Tạo và lưu dữ liệu training mới
            training_file = generate_training_data()
            print(f"Đã tạo dữ liệu training mới: {training_file}")
            # Đổi tên file mới nhất thành training_data.csv
            if training_file != 'training_data.csv':
                if os.path.exists('training_data.csv'):
                    os.remove('training_data.csv')
                os.rename(training_file, 'training_data.csv')
            ai.train()
    except Exception as e:
        print(f"Lỗi khi tải mô hình: {e}")
        print("Đang tạo mô hình mới...")
        # Tạo và lưu dữ liệu training mới
        training_file = generate_training_data()
        print(f"Đã tạo dữ liệu training mới: {training_file}")
        # Đổi tên file mới nhất thành training_data.csv
        if training_file != 'training_data.csv':
            if os.path.exists('training_data.csv'):
                os.remove('training_data.csv')
            os.rename(training_file, 'training_data.csv')
        ai.train()
else:
    print("Không tìm thấy file mô hình, đang tạo mới...")
    # Tạo và lưu dữ liệu training mới
    training_file = generate_training_data()
    print(f"Đã tạo dữ liệu training mới: {training_file}")
    # Đổi tên file mới nhất thành training_data.csv
    if training_file != 'training_data.csv':
        if os.path.exists('training_data.csv'):
            os.remove('training_data.csv')
        os.rename(training_file, 'training_data.csv')
    ai.train()

print("Khởi tạo mô hình AI hoàn tất!")

# ✅ Khởi tạo file Excel nếu chưa tồn tại
def init_excel():
    try:
        if not os.path.exists(USERS_EXCEL):
            wb = Workbook()
            ws = wb.active
            ws.title = "Users"
            # Thêm header
            headers = ['student_id', 'full_name', 'email', 'password', 'dob', 'join_date', 'image_url']
            ws.append(headers)
            wb.save(USERS_EXCEL)
            print(f"Đã tạo file Excel mới: {USERS_EXCEL}")
        else:
            # Kiểm tra xem file có phải là file Excel hợp lệ không
            try:
                wb = load_workbook(USERS_EXCEL)
                print(f"File Excel đã tồn tại và hợp lệ: {USERS_EXCEL}")
            except Exception as e:
                print(f"File Excel không hợp lệ, đang tạo lại: {e}")
                os.remove(USERS_EXCEL)
                wb = Workbook()
                ws = wb.active
                ws.title = "Users"
                headers = ['student_id', 'full_name', 'email', 'password', 'dob', 'join_date', 'image_url']
                ws.append(headers)
                wb.save(USERS_EXCEL)
                print(f"Đã tạo lại file Excel: {USERS_EXCEL}")
    except Exception as e:
        print(f"Lỗi khi khởi tạo Excel: {e}")
        raise

# ✅ Đọc người dùng từ Excel
def read_excel():
    try:
        wb = load_workbook(USERS_EXCEL)
        ws = wb.active
        users = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Kiểm tra nếu có dữ liệu
                users.append({
                    'student_id': row[0],
                    'full_name': row[1],
                    'email': row[2],
                    'password': row[3],
                    'dob': row[4],
                    'join_date': row[5],
                    'image_url': row[6]
                })
        return users
    except Exception as e:
        print(f"Lỗi khi đọc file Excel: {e}")
        return []

# ✅ Thêm người dùng mới
def add_user(name, dob, email, password):
    try:
        wb = load_workbook(USERS_EXCEL)
        ws = wb.active
        student_id = generate_student_id()
        join_date = datetime.now().strftime('%Y-%m-%d')
        ws.append([student_id, name, email, password, dob, join_date, None])
        wb.save(USERS_EXCEL)
        print(f"Đã thêm người dùng mới: {email}")
    except Exception as e:
        print(f"Lỗi khi thêm người dùng: {e}")
        raise

# ✅ Tạo mã học sinh mới
def generate_student_id():
    users = read_excel()
    if not users:
        return 'HS0001'
    last_id = max(user['student_id'] for user in users)
    num = int(last_id[2:]) + 1
    return f'HS{num:04d}'

# ✅ API đăng ký
@app.route("/api/register", methods=["POST"])
def register_user():
    try:
        data = request.get_json()
        users = read_excel()
        if any(user["email"] == data["email"] for user in users):
            return jsonify({"success": False, "message": "Email đã tồn tại."}), 400
        add_user(data["name"], data["dob"], data["email"], data["password"])
        return jsonify({"success": True})
    except Exception as e:
        print(f"Lỗi khi đăng ký: {e}")
        return jsonify({"success": False, "message": "Có lỗi xảy ra khi đăng ký."}), 500

# ✅ API đăng nhập
@app.route("/api/login", methods=["POST"])
def do_login():
    try:
        data = request.get_json()
        users = read_excel()
        for user in users:
            if user["email"] == data["email"] and user["password"] == data["password"]:
                return jsonify({"success": True})
        return jsonify({"success": False, "message": "Sai email hoặc mật khẩu."}), 401
    except Exception as e:
        print(f"Lỗi khi đăng nhập: {e}")
        return jsonify({"success": False, "message": "Có lỗi xảy ra khi đăng nhập."}), 500

# ✅ API lấy thông tin hồ sơ
@app.route("/api/profile", methods=["GET"])
def get_profile():
    try:
        # Lấy email từ query params hoặc session
        email = request.args.get("email")
        if not email:
            # Nếu không có email trong query params, trả về thông tin mặc định
            return jsonify({
                "success": True,
                "profile": {
                    "student_id": "",
                    "full_name": "",
                    "email": "",
                    "join_date": "",
                    "image_url": "/static/default-avatar.svg"
                },
                "notifications": {
                    "email": True,
                    "new_messages": True,
                    "assignment_updates": True,
                    "system_announcements": True
                }
            })

        users = read_excel()
        user = next((u for u in users if u["email"] == email), None)
        
        if not user:
            return jsonify({"success": False, "message": "Không tìm thấy thông tin người dùng"}), 404
        
        return jsonify({
            "success": True,
            "profile": {
                "student_id": user["student_id"],
                "full_name": user["full_name"],
                "email": user["email"],
                "join_date": user["join_date"],
                "image_url": user["image_url"] or "/static/default-avatar.svg"
            },
            "notifications": {
                "email": True,
                "new_messages": True,
                "assignment_updates": True,
                "system_announcements": True
            }
        })
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

# ✅ API cập nhật thông tin người dùng
@app.route('/api/update-profile', methods=['POST'])
def update_profile():
    try:
        # Lấy dữ liệu từ form
        full_name = request.form.get('full_name')
        email = request.form.get('email')
        student_id = request.form.get('student_id')
        
        # Kiểm tra email có tồn tại không
        if not email:
            return jsonify({'success': False, 'message': 'Email không được để trống'})

        # Đọc dữ liệu từ Excel
        wb = load_workbook(USERS_EXCEL)
        ws = wb.active
        
        # Tìm người dùng theo email
        user_found = False
        for row in ws.iter_rows(min_row=2):
            if row[2].value == email:  # Email ở cột thứ 3
                user_found = True
                # Cập nhật thông tin
                row[1].value = full_name  # Tên ở cột thứ 2
                
                # Xử lý upload ảnh nếu có
                if 'image' in request.files:
                    image = request.files['image']
                    if image and image.filename:
                        # Tạo tên file an toàn
                        filename = secure_filename(image.filename)
                        # Thêm timestamp để tránh trùng tên
                        filename = f"{int(time.time())}_{filename}"
                        # Lưu file
                        image_path = os.path.join('static/uploads', filename)
                        image.save(image_path)
                        # Cập nhật đường dẫn ảnh trong Excel
                        row[6].value = f'/static/uploads/{filename}'  # Image URL ở cột thứ 7
                
                # Lưu lại vào Excel
                wb.save(USERS_EXCEL)
                
                return jsonify({
                    'success': True,
                    'message': 'Cập nhật hồ sơ thành công',
                    'profile': {
                        'full_name': full_name,
                        'email': email,
                        'student_id': student_id,
                        'image_url': row[6].value or '/static/default-avatar.svg'
                    }
                })
        
        if not user_found:
            return jsonify({'success': False, 'message': 'Không tìm thấy người dùng'})
        
    except Exception as e:
        print(f"Error updating profile: {str(e)}")
        return jsonify({'success': False, 'message': 'Có lỗi xảy ra khi cập nhật hồ sơ'})

# ✅ API đổi mật khẩu
@app.route("/api/change-password", methods=["POST"])
def change_password():
    try:
        data = request.get_json()
        email = data.get("email")
        current_password = data.get("current_password")
        new_password = data.get("new_password")

        if not all([email, current_password, new_password]):
            return jsonify({"success": False, "message": "Vui lòng điền đầy đủ thông tin"}), 400

        wb = load_workbook(USERS_EXCEL)
        ws = wb.active
        
        # Tìm dòng chứa email và kiểm tra mật khẩu hiện tại
        for row in ws.iter_rows(min_row=2):
            if row[2].value == email:  # Email ở cột thứ 3
                if row[3].value != current_password:  # Password ở cột thứ 4
                    return jsonify({"success": False, "message": "Mật khẩu hiện tại không đúng"}), 400
                
                # Cập nhật mật khẩu mới
                row[3].value = new_password
                wb.save(USERS_EXCEL)
                return jsonify({"success": True, "message": "Đổi mật khẩu thành công"}), 200
        
        return jsonify({"success": False, "message": "Không tìm thấy thông tin người dùng"}), 404
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

# ✅ Trang chính
@app.route("/")
def home():
    return redirect("/login")

# ✅ Trang đăng nhập
@app.route("/login")
def login_page():
    return render_template("login.html")

# ✅ Trang đăng ký
@app.route("/register")
def register_page():
    return render_template("register.html")

# ✅ Trang dashboard
@app.route("/dashboard")
def dashboard():
    return render_template("index.html")

@app.route('/api/generate-study-plan', methods=['POST'])
def generate_study_plan():
    try:
        print("\n[DEBUG] ====== Bắt đầu xử lý request generate-study-plan ======")
        
        # Lấy dữ liệu từ request
        data = request.get_json()
        print("[DEBUG] Received data:", data)
        
        # Kiểm tra dữ liệu đầu vào
        required_fields = ['subject', 'grade', 'current_score', 'target_score', 
                         'duration_weeks', 'daily_study_hours', 'learning_style']
        
        # In ra kiểu dữ liệu của các trường
        print("[DEBUG] Data types:", {k: type(v).__name__ for k, v in data.items()})
        
        # Kiểm tra các trường bắt buộc
        for field in required_fields:
            if field not in data:
                return jsonify({
                    'error': f'Thiếu trường {field}'
                }), 400
        
        # Chuyển đổi các giá trị số
        try:
            current_score = float(data['current_score'])
            target_score = float(data['target_score'])
            duration_weeks = int(data['duration_weeks'])
            daily_study_hours = float(data['daily_study_hours'])
            
            print(f"[DEBUG] Converted values: current_score={current_score} ({type(current_score)}), "
                  f"target_score={target_score} ({type(target_score)}), "
                  f"duration_weeks={duration_weeks} ({type(duration_weeks)}), "
                  f"daily_study_hours={daily_study_hours} ({type(daily_study_hours)})")
            
        except ValueError as e:
            return jsonify({
                'error': f'Giá trị không hợp lệ: {str(e)}'
            }), 400
        
        # Kiểm tra giá trị hợp lệ
        if not (0 <= current_score <= 10 and 0 <= target_score <= 10):
            return jsonify({
                'error': 'Điểm số phải nằm trong khoảng 0-10'
            }), 400
            
        if duration_weeks <= 0 or daily_study_hours <= 0:
            return jsonify({
                'error': 'Thời gian học phải lớn hơn 0'
            }), 400
        
        # Gọi AI để tạo lộ trình học tập
        print("[DEBUG] Calling AI generate_learning_path with:", {
            'subject': data['subject'],
            'current_score': current_score,
            'target_score': target_score,
            'duration_weeks': duration_weeks,
            'daily_study_hours': daily_study_hours,
            'learning_style': data['learning_style'],
            'grade': data['grade']
        })
        
        learning_path = ai.generate_learning_path(
            subject=data['subject'],
            current_score=current_score,
            target_score=target_score,
            duration_weeks=duration_weeks,
            daily_study_hours=daily_study_hours,
            learning_style=data['learning_style'],
            grade=data['grade']
        )
        
        if learning_path is None:
            print("[ERROR] AI returned None (không thể tạo lộ trình học tập)")
            return jsonify({
                'error': 'Không thể tạo lộ trình học tập. Vui lòng thử lại.'
            }), 500
            
        print("[DEBUG] Learning path generated successfully")
        print("[DEBUG] Learning path structure:", {
            'type': type(learning_path).__name__,
            'length': len(learning_path),
            'first_week': type(learning_path[0]).__name__ if learning_path else None
        })
        
        return jsonify({
            'success': True,
            'learning_path': learning_path
        })
        
    except Exception as e:
        print("\n[ERROR] Exception in generate_study_plan:")
        print(f"[ERROR] Error type: {type(e).__name__}")
        print(f"[ERROR] Error message: {str(e)}")
        import traceback
        print("[ERROR] Stack trace:")
        traceback.print_exc()
        return jsonify({
            'error': 'Đã xảy ra lỗi khi tạo lộ trình học tập'
        }), 500

@app.route('/favicon.ico')
def favicon():
    return send_from_directory(os.path.join(app.root_path, 'static'),
                             'favicon.ico', mimetype='image/vnd.microsoft.icon')

# Hàm quản lý Grade
def add_grade(name_grade):
    try:
        wb = load_workbook(GRADE_EXCEL)
        ws = wb.active
        grade_id = f"G{len(list(ws.rows)):03d}"
        ws.append([grade_id, name_grade])
        wb.save(GRADE_EXCEL)
        return grade_id
    except Exception as e:
        print(f"Lỗi khi thêm grade: {e}")
        raise

def get_all_grades():
    try:
        wb = load_workbook(GRADE_EXCEL)
        ws = wb.active
        grades = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                grades.append({
                    'ID_grade': row[0],
                    'Name_grade': row[1]
                })
        return grades
    except Exception as e:
        print(f"Lỗi khi đọc grades: {e}")
        return []

# Hàm quản lý Subject
def add_subject(name_subject, grade_id):
    try:
        wb = load_workbook(SUBJECT_EXCEL)
        ws = wb.active
        subject_id = f"S{len(list(ws.rows)):03d}"
        ws.append([subject_id, name_subject, grade_id])
        wb.save(SUBJECT_EXCEL)
        return subject_id
    except Exception as e:
        print(f"Lỗi khi thêm subject: {e}")
        raise

def get_subjects_by_grade(grade_id):
    try:
        wb = load_workbook(SUBJECT_EXCEL)
        ws = wb.active
        subjects = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[2] == grade_id:
                subjects.append({
                    'ID_subject': row[0],
                    'name_subject': row[1],
                    'ID_grade': row[2]
                })
        return subjects
    except Exception as e:
        print(f"Lỗi khi đọc subjects: {e}")
        return []

# Hàm quản lý Topic
def add_topic(topic_name, subject_id, grade_id):
    try:
        wb = load_workbook(TOPIC_EXCEL)
        ws = wb.active
        topic_id = f"T{len(list(ws.rows)):03d}"
        ws.append([topic_id, topic_name, subject_id, grade_id])
        wb.save(TOPIC_EXCEL)
        return topic_id
    except Exception as e:
        print(f"Lỗi khi thêm topic: {e}")
        raise

def get_topics_by_subject(subject_id):
    try:
        wb = load_workbook(TOPIC_EXCEL)
        ws = wb.active
        topics = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[2] == subject_id:
                topics.append({
                    'ID_topic': row[0],
                    'topic_name': row[1],
                    'ID_subject': row[2],
                    'ID_grade': row[3]
                })
        return topics
    except Exception as e:
        print(f"Lỗi khi đọc topics: {e}")
        return []

# Hàm quản lý Theory
def add_theory(theory_name, topic_id, subject_id, grade_id, level, url, completion_time):
    try:
        wb = load_workbook(THEORY_EXCEL)
        ws = wb.active
        theory_id = f"TH{len(list(ws.rows)):03d}"
        ws.append([theory_id, theory_name, topic_id, subject_id, grade_id, level, url, completion_time])
        wb.save(THEORY_EXCEL)
        return theory_id
    except Exception as e:
        print(f"Lỗi khi thêm theory: {e}")
        raise

def get_theories_by_topic(topic_id):
    try:
        wb = load_workbook(THEORY_EXCEL)
        ws = wb.active
        theories = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[2] == topic_id:
                theories.append({
                    'ID_theory': row[0],
                    'theory_name': row[1],
                    'ID_topic': row[2],
                    'ID_subject': row[3],
                    'ID_grade': row[4],
                    'level': row[5],
                    'URL': row[6],
                    'completion_time': row[7]
                })
        return theories
    except Exception as e:
        print(f"Lỗi khi đọc theories: {e}")
        return []

# Hàm quản lý Practice
def add_practice(practice_name, topic_id, subject_id, grade_id, level, theory_id):
    try:
        wb = load_workbook(PRACTICE_EXCEL)
        ws = wb.active
        practice_id = f"P{len(list(ws.rows)):03d}"
        ws.append([practice_id, practice_name, topic_id, subject_id, grade_id, level, theory_id])
        wb.save(PRACTICE_EXCEL)
        return practice_id
    except Exception as e:
        print(f"Lỗi khi thêm practice: {e}")
        raise

def get_practices_by_theory(theory_id):
    try:
        wb = load_workbook(PRACTICE_EXCEL)
        ws = wb.active
        practices = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[6] == theory_id:
                practices.append({
                    'ID_practice': row[0],
                    'practice_name': row[1],
                    'ID_topic': row[2],
                    'ID_subject': row[3],
                    'ID_grade': row[4],
                    'level': row[5],
                    'ID_theory': row[6]
                })
        return practices
    except Exception as e:
        print(f"Lỗi khi đọc practices: {e}")
        return []

# API endpoints cho quản lý dữ liệu
@app.route('/api/grades', methods=['GET'])
def get_grades():
    try:
        grades = get_all_grades()
        return jsonify({"success": True, "grades": grades})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/subjects/<grade_id>', methods=['GET'])
def get_subjects(grade_id):
    try:
        subjects = get_subjects_by_grade(grade_id)
        return jsonify({"success": True, "subjects": subjects})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/topics/<subject_id>', methods=['GET'])
def get_topics(subject_id):
    try:
        topics = get_topics_by_subject(subject_id)
        return jsonify({"success": True, "topics": topics})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/theories/<topic_id>', methods=['GET'])
def get_theories(topic_id):
    try:
        theories = get_theories_by_topic(topic_id)
        return jsonify({"success": True, "theories": theories})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/practices/<theory_id>', methods=['GET'])
def get_practices(theory_id):
    try:
        practices = get_practices_by_theory(theory_id)
        return jsonify({"success": True, "practices": practices})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

# API endpoints cho thêm dữ liệu mới
@app.route('/api/grades', methods=['POST'])
def create_grade():
    try:
        data = request.get_json()
        if not data or 'name_grade' not in data:
            return jsonify({"success": False, "message": "Thiếu thông tin name_grade"}), 400
        
        grade_id = add_grade(data['name_grade'])
        return jsonify({"success": True, "grade_id": grade_id})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/subjects', methods=['POST'])
def create_subject():
    try:
        data = request.get_json()
        if not data or 'name_subject' not in data or 'grade_id' not in data:
            return jsonify({"success": False, "message": "Thiếu thông tin name_subject hoặc grade_id"}), 400
        
        subject_id = add_subject(data['name_subject'], data['grade_id'])
        return jsonify({"success": True, "subject_id": subject_id})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/topics', methods=['POST'])
def create_topic():
    try:
        data = request.get_json()
        if not data or 'topic_name' not in data or 'subject_id' not in data or 'grade_id' not in data:
            return jsonify({"success": False, "message": "Thiếu thông tin topic_name, subject_id hoặc grade_id"}), 400
        
        topic_id = add_topic(data['topic_name'], data['subject_id'], data['grade_id'])
        return jsonify({"success": True, "topic_id": topic_id})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/theories', methods=['POST'])
def create_theory():
    try:
        data = request.get_json()
        required_fields = ['theory_name', 'topic_id', 'subject_id', 'grade_id', 'level', 'url', 'completion_time']
        if not data or not all(field in data for field in required_fields):
            return jsonify({"success": False, "message": "Thiếu thông tin bắt buộc"}), 400
        
        theory_id = add_theory(
            data['theory_name'],
            data['topic_id'],
            data['subject_id'],
            data['grade_id'],
            data['level'],
            data['url'],
            data['completion_time']
        )
        return jsonify({"success": True, "theory_id": theory_id})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/practices', methods=['POST'])
def create_practice():
    try:
        data = request.get_json()
        required_fields = ['practice_name', 'topic_id', 'subject_id', 'grade_id', 'level', 'theory_id']
        if not data or not all(field in data for field in required_fields):
            return jsonify({"success": False, "message": "Thiếu thông tin bắt buộc"}), 400
        
        practice_id = add_practice(
            data['practice_name'],
            data['topic_id'],
            data['subject_id'],
            data['grade_id'],
            data['level'],
            data['theory_id']
        )
        return jsonify({"success": True, "practice_id": practice_id})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

def init_sample_data():
    """Khởi tạo dữ liệu mẫu cho các file Excel"""
    try:
        # Thêm các lớp học
        grades = [
            "Lớp 10",
            "Lớp 11",
            "Lớp 12"
        ]
        grade_ids = {}
        for grade in grades:
            grade_id = add_grade(grade)
            grade_ids[grade] = grade_id

        # Thêm các môn học
        subjects = {
            "Lớp 10": ["Toán", "Vật Lý", "Hóa Học"],
            "Lớp 11": ["Toán", "Vật Lý", "Hóa Học"],
            "Lớp 12": ["Toán", "Vật Lý", "Hóa Học"]
        }
        subject_ids = {}
        for grade, subject_list in subjects.items():
            grade_id = grade_ids[grade]
            for subject in subject_list:
                subject_id = add_subject(subject, grade_id)
                subject_ids[f"{grade}_{subject}"] = subject_id

        # Thêm các chủ đề cho Toán lớp 10
        math_10_topics = [
            "Mệnh đề – Tập hợp",
            "Hàm số bậc nhất và bậc hai",
            "Thống kê và xác suất",
            "Phương pháp tọa độ trong mặt phẳng",
            "Hình học không gian"
        ]
        topic_ids = {}
        for topic in math_10_topics:
            topic_id = add_topic(topic, subject_ids["Lớp 10_Toán"], grade_ids["Lớp 10"])
            topic_ids[f"Lớp 10_Toán_{topic}"] = topic_id

        # Thêm bài học lý thuyết mẫu cho chủ đề "Mệnh đề – Tập hợp"
        theory_lessons = [
            {
                "name": "Khái niệm mệnh đề, phủ định, mệnh đề kéo theo, mệnh đề đảo",
                "level": "basic",
                "url": "https://example.com/lesson1",
                "completion_time": 45
            },
            {
                "name": "Tập hợp, phần tử, các phép toán trên tập hợp",
                "level": "basic",
                "url": "https://example.com/lesson2",
                "completion_time": 60
            }
        ]
        theory_ids = {}
        for lesson in theory_lessons:
            theory_id = add_theory(
                lesson["name"],
                topic_ids["Lớp 10_Toán_Mệnh đề – Tập hợp"],
                subject_ids["Lớp 10_Toán"],
                grade_ids["Lớp 10"],
                lesson["level"],
                lesson["url"],
                lesson["completion_time"]
            )
            theory_ids[lesson["name"]] = theory_id

        print("Đã khởi tạo dữ liệu mẫu thành công!")
        return True
    except Exception as e:
        print(f"Lỗi khi khởi tạo dữ liệu mẫu: {e}")
        return False

# Gọi hàm khởi tạo dữ liệu mẫu khi khởi động ứng dụng
init_sample_data()

def generate_practice_exercises(theory_id):
    """Tạo bài tập thực hành dựa trên bài học lý thuyết sử dụng AI"""
    try:
        # Lấy thông tin bài học lý thuyết
        wb = load_workbook(THEORY_EXCEL)
        ws = wb.active
        theory_info = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == theory_id:
                theory_info = {
                    'ID_theory': row[0],
                    'theory_name': row[1],
                    'ID_topic': row[2],
                    'ID_subject': row[3],
                    'ID_grade': row[4],
                    'level': row[5],
                    'URL': row[6],
                    'completion_time': row[7]
                }
                break
        
        if not theory_info:
            raise Exception("Không tìm thấy bài học lý thuyết")

        # Lấy thông tin chủ đề
        wb = load_workbook(TOPIC_EXCEL)
        ws = wb.active
        topic_info = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == theory_info['ID_topic']:
                topic_info = {
                    'ID_topic': row[0],
                    'topic_name': row[1],
                    'ID_subject': row[2],
                    'ID_grade': row[3]
                }
                break

        if not topic_info:
            raise Exception("Không tìm thấy thông tin chủ đề")

        # Tạo bài tập thực hành sử dụng AI
        # TODO: Thay thế phần này bằng việc gọi API của mô hình AI thực tế
        practice_exercises = [
            {
                "practice_name": f"Bài tập 1: {theory_info['theory_name']}",
                "level": theory_info['level'],
                "content": "Nội dung bài tập sẽ được tạo bởi AI"
            },
            {
                "practice_name": f"Bài tập 2: {theory_info['theory_name']}",
                "level": theory_info['level'],
                "content": "Nội dung bài tập sẽ được tạo bởi AI"
            },
            {
                "practice_name": f"Bài tập 3: {theory_info['theory_name']}",
                "level": theory_info['level'],
                "content": "Nội dung bài tập sẽ được tạo bởi AI"
            }
        ]

        # Thêm các bài tập vào database
        practice_ids = []
        for exercise in practice_exercises:
            practice_id = add_practice(
                exercise["practice_name"],
                theory_info['ID_topic'],
                theory_info['ID_subject'],
                theory_info['ID_grade'],
                exercise["level"],
                theory_id
            )
            practice_ids.append(practice_id)

        return practice_ids
    except Exception as e:
        print(f"Lỗi khi tạo bài tập thực hành: {e}")
        raise

@app.route('/api/generate-practice/<theory_id>', methods=['POST'])
def generate_practice(theory_id):
    """API endpoint để tạo bài tập thực hành cho một bài học lý thuyết"""
    try:
        practice_ids = generate_practice_exercises(theory_id)
        return jsonify({
            "success": True,
            "message": "Đã tạo bài tập thực hành thành công",
            "practice_ids": practice_ids
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Lỗi khi tạo bài tập thực hành: {str(e)}"
        }), 500

# ✅ Khởi chạy app
if __name__ == "__main__":
    init_excel()
    # Tạo và lưu dữ liệu training mới khi khởi động app
    training_file = generate_training_data()
    print(f"Đã tạo dữ liệu training mới: {training_file}")
    # Đổi tên file mới nhất thành training_data.csv
    if training_file != 'training_data.csv':
        if os.path.exists('training_data.csv'):
            os.remove('training_data.csv')
        os.rename(training_file, 'training_data.csv')
    app.run(debug=True)
